#!/usr/bin/env python3
"""
Generate a George Washington scatterplot from Excel data.

Expected Excel columns:
- X_coordinate
- Y_coordinate
- Intensity

Default input:
  /Users/zane/Desktop/george_washington/george_washington_v1/george_washington_v1_input.xlsx

Default output:
  /Users/zane/Desktop/george_washington/george_washington_v1/george_washington_v1_scatterplot.png

Default Excel output:
  /Users/zane/Desktop/george_washington/george_washington_v1/george_washington_v1_with_data.xlsx

Default interactive 3D output:
  /Users/zane/Desktop/george_washington/george_washington_v1/george_washington_v1_3d_plot.html
"""

from __future__ import annotations

import argparse
import os
import tempfile
from pathlib import Path
from typing import Any


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = SCRIPT_DIR / "george_washington_v1_input.xlsx"
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "george_washington_v1_scatterplot.png"
DEFAULT_EXCEL_OUTPUT_PATH = SCRIPT_DIR / "george_washington_v1_with_data.xlsx"
DEFAULT_HTML_OUTPUT_PATH = SCRIPT_DIR / "george_washington_v1_3d_plot.html"
DEFAULT_MATPLOTLIB_CACHE = SCRIPT_DIR / ".matplotlib_cache"
REQUIRED_COLUMNS = ["X_coordinate", "Y_coordinate", "Intensity"]


def configure_matplotlib_cache() -> None:
    if "MPLCONFIGDIR" in os.environ:
        return

    try:
        DEFAULT_MATPLOTLIB_CACHE.mkdir(parents=True, exist_ok=True)
        os.environ["MPLCONFIGDIR"] = str(DEFAULT_MATPLOTLIB_CACHE)
    except OSError:
        os.environ["MPLCONFIGDIR"] = tempfile.mkdtemp(prefix="matplotlib-cache-")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a George Washington face scatterplot from Excel data."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Input Excel file. Default: {DEFAULT_EXCEL_PATH}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output PNG path. Default: {DEFAULT_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=DEFAULT_EXCEL_OUTPUT_PATH,
        help=f"Output Excel workbook path. Default: {DEFAULT_EXCEL_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--html-output",
        type=Path,
        default=DEFAULT_HTML_OUTPUT_PATH,
        help=f"Output interactive 3D HTML path. Default: {DEFAULT_HTML_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="Open an interactive plot window after saving the PNG.",
    )
    parser.add_argument(
        "--point-size",
        type=float,
        default=1.0,
        help="Scatter point size. Default: 1.0",
    )
    parser.add_argument(
        "--alpha",
        type=float,
        default=0.7,
        help="Point transparency from 0 to 1. Default: 0.7",
    )
    return parser.parse_args()


def load_scatter_data(excel_path: Path) -> Any:
    import pandas as pd

    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    scatter_df = pd.read_excel(excel_path)
    missing_columns = [
        column for column in REQUIRED_COLUMNS if column not in scatter_df.columns
    ]
    if missing_columns:
        raise ValueError(
            "Excel file is missing required columns: "
            + ", ".join(missing_columns)
        )

    return scatter_df


def create_scatterplot(
    scatter_df: Any,
    output_path: Path,
    point_size: float,
    alpha: float,
    show_plot: bool,
) -> None:
    configure_matplotlib_cache()

    import matplotlib

    if not show_plot:
        matplotlib.use("Agg")

    import matplotlib.pyplot as plt

    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_x = scatter_df["X_coordinate"].values
    plot_y = scatter_df["Y_coordinate"].values
    plot_c = scatter_df["Intensity"].values

    plt.figure(figsize=(10, 10))
    plt.scatter(
        plot_x,
        plot_y,
        c=plot_c,
        cmap="viridis_r",
        s=point_size,
        alpha=alpha,
    )
    plt.title("George Washington's Face as a Scatter Plot (from Excel Data)")
    plt.xlabel("X-coordinate")
    plt.ylabel("Y-coordinate")
    plt.gca().set_aspect("equal", adjustable="box")

    plt.text(
        5,
        5,
        "George Washington",
        fontsize=12,
        color="black",
        bbox=dict(facecolor="white", alpha=0.7, edgecolor="none"),
    )

    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    print(f"Saved scatterplot to {output_path}")

    if show_plot:
        plt.show()
    else:
        plt.close()


def write_excel_with_plot(
    scatter_df: Any,
    plot_path: Path,
    excel_output_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    plot_path = plot_path.expanduser().resolve()
    excel_output_path = excel_output_path.expanduser().resolve()
    excel_output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    data_ws = wb.active
    data_ws.title = "Data"
    for row in dataframe_to_rows(scatter_df, index=False, header=True):
        data_ws.append(row)

    header_fill = PatternFill("solid", start_color="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in data_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = data_ws.dimensions
    data_ws.column_dimensions["A"].width = 16
    data_ws.column_dimensions["B"].width = 16
    data_ws.column_dimensions["C"].width = 14

    plot_ws = wb.create_sheet("Scatterplot")
    plot_ws["A1"] = "George Washington Scatterplot"
    plot_ws["A1"].font = Font(bold=True, size=16)

    image = ExcelImage(str(plot_path))
    image.width = 720
    image.height = 720
    plot_ws.add_image(image, "A3")

    wb.save(excel_output_path)
    print(f"Saved Excel workbook to {excel_output_path}")


def create_interactive_3d_plot(
    scatter_df: Any,
    html_output_path: Path,
    show_plot: bool,
) -> None:
    try:
        import plotly.express as px
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "Plotly is required for the interactive 3D HTML plot. "
            "Install it with:\n"
            "  python3 -m pip install plotly"
        ) from exc

    html_output_path = html_output_path.expanduser().resolve()
    html_output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_x = scatter_df["X_coordinate"].values
    plot_y = scatter_df["Y_coordinate"].values
    plot_c = scatter_df["Intensity"].values

    fig = px.scatter_3d(
        x=plot_x,
        y=plot_y,
        z=plot_c,
        color=plot_c,
        color_continuous_scale="rdbu",
        title="Interactive 3D Scatter Plot of George Washington's Face (Plotly)",
        labels={
            "x": "X-coordinate",
            "y": "Y-coordinate",
            "z": "Intensity",
            "color": "Intensity",
        },
    )
    fig.update_traces(marker=dict(size=2, opacity=0.7))
    fig.write_html(html_output_path)
    print(f"Saved interactive 3D plot to {html_output_path}")

    if show_plot:
        fig.show()


def main() -> None:
    args = parse_args()
    scatter_df = load_scatter_data(args.excel)
    output_path = args.output.expanduser().resolve()

    create_scatterplot(
        scatter_df=scatter_df,
        output_path=output_path,
        point_size=args.point_size,
        alpha=args.alpha,
        show_plot=args.show,
    )
    write_excel_with_plot(
        scatter_df=scatter_df,
        plot_path=output_path,
        excel_output_path=args.excel_output,
    )
    create_interactive_3d_plot(
        scatter_df=scatter_df,
        html_output_path=args.html_output,
        show_plot=args.show,
    )


if __name__ == "__main__":
    main()
