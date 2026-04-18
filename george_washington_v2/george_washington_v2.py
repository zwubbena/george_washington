#!/usr/bin/env python3
"""
Generate a scatterplot PNG and Excel workbook from george_washington_v2_input.xlsx.

The script reads the Data sheet from:
  /Users/zane/Desktop/george_washington/george_washington_v2/george_washington_v2_input.xlsx

Expected Data sheet columns:
- x
- y

Default outputs:
- george_washington_v2_scatterplot.png
- george_washington_v2_with_data.xlsx
"""

from __future__ import annotations

import argparse
import os
import tempfile
from pathlib import Path
from typing import Any


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_WORKBOOK = SCRIPT_DIR / "george_washington_v2_input.xlsx"
DEFAULT_OUTPUT_PNG = SCRIPT_DIR / "george_washington_v2_scatterplot.png"
DEFAULT_OUTPUT_WORKBOOK = SCRIPT_DIR / "george_washington_v2_with_data.xlsx"
DEFAULT_MATPLOTLIB_CACHE = SCRIPT_DIR / ".matplotlib_cache"
DATA_SHEET_NAME = "Data"
REQUIRED_COLUMNS = ["x", "y"]


def configure_matplotlib_cache() -> None:
    if "MPLCONFIGDIR" in os.environ:
        return

    try:
        DEFAULT_MATPLOTLIB_CACHE.mkdir(parents=True, exist_ok=True)
        os.environ["MPLCONFIGDIR"] = str(DEFAULT_MATPLOTLIB_CACHE)
    except OSError:
        os.environ["MPLCONFIGDIR"] = tempfile.mkdtemp(prefix="matplotlib-cache-")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a scatterplot PNG and Excel workbook from george_washington_v2_input.xlsx."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_WORKBOOK,
        help=f"Input workbook. Default: {DEFAULT_INPUT_WORKBOOK}",
    )
    parser.add_argument(
        "--sheet",
        default=DATA_SHEET_NAME,
        help=f"Worksheet to read. Default: {DATA_SHEET_NAME}",
    )
    parser.add_argument(
        "--png-output",
        type=Path,
        default=DEFAULT_OUTPUT_PNG,
        help=f"Output PNG path. Default: {DEFAULT_OUTPUT_PNG}",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=DEFAULT_OUTPUT_WORKBOOK,
        help=f"Output Excel workbook path. Default: {DEFAULT_OUTPUT_WORKBOOK}",
    )
    parser.add_argument(
        "--title",
        default="Washington Scatterplot",
        help="Scatterplot title. Default: Washington Scatterplot",
    )
    parser.add_argument(
        "--point-size",
        type=float,
        default=1.0,
        help="Scatter point size. Default: 1.0",
    )
    parser.add_argument(
        "--alpha",
        type=float,
        default=0.7,
        help="Point transparency from 0 to 1. Default: 0.7",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="Open an interactive plot window after saving the PNG.",
    )
    return parser.parse_args()


def load_data(workbook_path: Path, sheet_name: str) -> Any:
    import pandas as pd

    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {workbook_path}")

    scatter_df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    missing_columns = [
        column for column in REQUIRED_COLUMNS if column not in scatter_df.columns
    ]
    if missing_columns:
        raise ValueError(
            f"Worksheet '{sheet_name}' is missing required columns: "
            + ", ".join(missing_columns)
        )

    return scatter_df


def create_png_scatterplot(
    scatter_df: Any,
    png_output_path: Path,
    title: str,
    point_size: float,
    alpha: float,
    show_plot: bool,
) -> None:
    configure_matplotlib_cache()

    import matplotlib

    if not show_plot:
        matplotlib.use("Agg")

    import matplotlib.pyplot as plt

    png_output_path = png_output_path.expanduser().resolve()
    png_output_path.parent.mkdir(parents=True, exist_ok=True)

    plt.figure(figsize=(10, 10))
    plt.scatter(
        scatter_df["x"].values,
        scatter_df["y"].values,
        s=point_size,
        alpha=alpha,
        color="black",
    )
    plt.title(title)
    plt.xlabel("x")
    plt.ylabel("y")
    plt.gca().set_aspect("equal", adjustable="box")
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.savefig(png_output_path, dpi=300)
    print(f"Saved scatterplot PNG to {png_output_path}")

    if show_plot:
        plt.show()
    else:
        plt.close()


def write_excel_with_plot(
    scatter_df: Any,
    png_path: Path,
    excel_output_path: Path,
    title: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    png_path = png_path.expanduser().resolve()
    excel_output_path = excel_output_path.expanduser().resolve()
    excel_output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    data_ws = wb.active
    data_ws.title = "Data"
    for row in dataframe_to_rows(scatter_df, index=False, header=True):
        data_ws.append(row)

    header_fill = PatternFill("solid", start_color="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in data_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = data_ws.dimensions
    data_ws.column_dimensions["A"].width = 14
    data_ws.column_dimensions["B"].width = 14

    plot_ws = wb.create_sheet("Scatterplot")
    plot_ws["A1"] = title
    plot_ws["A1"].font = Font(bold=True, size=16)

    image = ExcelImage(str(png_path))
    image.width = 720
    image.height = 720
    plot_ws.add_image(image, "A3")

    wb.save(excel_output_path)
    print(f"Saved Excel workbook to {excel_output_path}")


def main() -> None:
    args = parse_args()
    scatter_df = load_data(args.input, args.sheet)
    png_output_path = args.png_output.expanduser().resolve()

    create_png_scatterplot(
        scatter_df=scatter_df,
        png_output_path=png_output_path,
        title=args.title,
        point_size=args.point_size,
        alpha=args.alpha,
        show_plot=args.show,
    )
    write_excel_with_plot(
        scatter_df=scatter_df,
        png_path=png_output_path,
        excel_output_path=args.excel_output,
        title=args.title,
    )


if __name__ == "__main__":
    main()
